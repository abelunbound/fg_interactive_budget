#!/usr/bin/env python3
"""
Consolidate spread-out Nigerian budget Excel data into clean table structure.
V4: Cell-level sequential parsing - robust to column misalignment.

USAGE:
    python3 consolidate_budget_v4.py

BEFORE RUNNING:
    Edit lines 207-208 to point to your files:
    input_file = '/path/to/your/full_budget_file.xlsx'
    output_file = '/path/to/output/consolidated_budget.xlsx'

Classification Logic:
1. CODE starts with "ERGP" → Project
   - Sequential parse: CODE, NAME, TYPE (optional text), AMOUNT (numeric)
2. CODE is 8+ digit number, no AMOUNT → MDA Header
   - Sequential parse: CODE, NAME
3. Everything else with AMOUNT → Budget Line
   - Sequential parse: CODE, LINE_ITEM, AMOUNT (numeric)

V4 Improvements:
- No fixed column positions - searches sequentially through row
- Resilient to column shifts and irregular spacing
- Handles missing TYPE values correctly
- Cell-level parsing (not character-level)
"""

import pandas as pd
from openpyxl import load_workbook


def is_numeric(value):
    """
    Check if a value is numeric (handles integers, floats, strings with numbers).
    """
    if value is None:
        return False
    
    val_str = str(value).strip().replace(',', '').replace('.', '', 1).replace('-', '', 1)
    return val_str.isdigit()


def extract_project_fields(cells):
    """
    Extract fields from ERGP project row using sequential parsing.
    
    Expected sequence: CODE, NAME, TYPE (optional), AMOUNT
    - TYPE is non-numeric text (e.g., "ONGOING", "NEW", or empty)
    - AMOUNT is numeric value
    
    Args:
        cells: List of non-empty cell values
    
    Returns:
        dict: Extracted fields (CODE, NAME, TYPE, AMOUNT)
    """
    if len(cells) < 2:
        return None
    
    code = str(cells[0]).strip()
    name = str(cells[1]).strip() if len(cells) > 1 else None
    type_val = None
    amount = None
    
    # Search remaining cells (from index 2 onwards)
    for i in range(2, len(cells)):
        cell_value = cells[i]
        
        if is_numeric(cell_value):
            # This is a numeric value
            if type_val is None:
                # Haven't found TYPE yet - keep searching
                # (This number might be a year, code, etc.)
                continue
            else:
                # TYPE already found - this is AMOUNT
                amount = cell_value
                break  # Stop once we have AMOUNT
        else:
            # Non-numeric text - this is TYPE
            if type_val is None:
                type_val = str(cell_value).strip()
    
    # Handle case where TYPE is missing but AMOUNT exists
    # If we haven't found AMOUNT yet and TYPE is still None,
    # take the first numeric value as AMOUNT
    if amount is None and type_val is None:
        for i in range(2, len(cells)):
            if is_numeric(cells[i]):
                amount = cells[i]
                break
    
    return {
        'CODE': code,
        'NAME': name,
        'TYPE': type_val,
        'AMOUNT': amount
    }


def extract_budget_line_fields(cells):
    """
    Extract fields from budget line row using sequential parsing.
    
    Expected sequence: CODE, LINE_ITEM, AMOUNT
    
    Args:
        cells: List of non-empty cell values
    
    Returns:
        dict: Extracted fields (CODE, LINE_ITEM, AMOUNT)
    """
    if len(cells) < 2:
        return None
    
    code = str(cells[0]).strip()
    line_item = str(cells[1]).strip() if len(cells) > 1 else None
    amount = None
    
    # Search for first numeric value as AMOUNT
    for i in range(2, len(cells)):
        if is_numeric(cells[i]):
            amount = cells[i]
            break
    
    return {
        'CODE': code,
        'LINE_ITEM': line_item,
        'AMOUNT': amount
    }


def extract_mda_header_fields(cells):
    """
    Extract fields from MDA header row.
    
    Expected sequence: CODE, NAME
    
    Args:
        cells: List of non-empty cell values
    
    Returns:
        dict: Extracted fields (CODE, NAME)
    """
    if len(cells) < 2:
        return None
    
    code = str(cells[0]).strip()
    name = str(cells[1]).strip() if len(cells) > 1 else None
    
    return {
        'CODE': code,
        'NAME': name
    }


def consolidate_budget_excel_v4(input_file, output_file):
    """
    Consolidate budget data using cell-level sequential parsing.
    
    Args:
        input_file: Path to input Excel file (the messy multi-column file)
        output_file: Path to output Excel file (clean multi-sheet file)
    
    Returns:
        tuple: (df_projects, df_budget_lines, df_all)
    """
    
    print(f"Reading: {input_file}")
    print("This may take a few minutes for large files...")
    
    # Load workbook
    wb = load_workbook(input_file, read_only=True)
    ws = wb.active
    
    # Collect cleaned data
    cleaned_rows = []
    current_mda_code = None
    current_mda_name = None
    
    print("Processing rows...")
    row_count = 0
    
    for row_num, row in enumerate(ws.iter_rows(values_only=True), 1):
        # Extract all non-empty cells from the row
        cells = [cell for cell in row if cell is not None and str(cell).strip() != ""]
        
        # Skip completely empty rows
        if len(cells) == 0:
            continue
        
        # Get the first cell as CODE
        code = str(cells[0]).strip() if len(cells) > 0 else ""
        
        # Skip header rows
        if code == "CODE" and len(cells) > 1:
            second_cell = str(cells[1]).strip()
            if second_cell in ["PROJECT NAME", "LINE ITEM"]:
                continue
        
        # Classify row based on CODE pattern
        
        # 1. ERGP CODE → Project
        if code.startswith("ERGP"):
            fields = extract_project_fields(cells)
            if fields:
                cleaned_rows.append({
                    'ROW_TYPE': 'PROJECT',
                    'MDA_CODE': current_mda_code,
                    'MDA_NAME': current_mda_name,
                    'CODE': fields['CODE'],
                    'ITEM_NAME': fields['NAME'],
                    'TYPE': fields['TYPE'],
                    'AMOUNT': fields['AMOUNT']
                })
                row_count += 1
        
        # 2. 8+ DIGIT CODE without numeric values in remaining cells → MDA Header
        elif code.isdigit() and len(code) >= 8:
            # Check if there are any numeric values in remaining cells
            has_amount = any(is_numeric(cell) for cell in cells[1:])
            
            if not has_amount:
                # This is an MDA header
                fields = extract_mda_header_fields(cells)
                if fields:
                    current_mda_code = fields['CODE']
                    current_mda_name = fields['NAME']
                    
                    cleaned_rows.append({
                        'ROW_TYPE': 'MDA_HEADER',
                        'MDA_CODE': fields['CODE'],
                        'MDA_NAME': fields['NAME'],
                        'CODE': None,
                        'ITEM_NAME': None,
                        'TYPE': None,
                        'AMOUNT': None
                    })
                    row_count += 1
            else:
                # This is a budget line (8+ digit code with amount)
                fields = extract_budget_line_fields(cells)
                if fields:
                    cleaned_rows.append({
                        'ROW_TYPE': 'BUDGET_LINE',
                        'MDA_CODE': current_mda_code,
                        'MDA_NAME': current_mda_name,
                        'CODE': fields['CODE'],
                        'ITEM_NAME': fields['LINE_ITEM'],
                        'TYPE': None,
                        'AMOUNT': fields['AMOUNT']
                    })
                    row_count += 1
        
        # 3. Everything else with numeric values → Budget Line
        elif len(cells) >= 2:
            has_amount = any(is_numeric(cell) for cell in cells[1:])
            
            if has_amount:
                fields = extract_budget_line_fields(cells)
                if fields:
                    cleaned_rows.append({
                        'ROW_TYPE': 'BUDGET_LINE',
                        'MDA_CODE': current_mda_code,
                        'MDA_NAME': current_mda_name,
                        'CODE': fields['CODE'],
                        'ITEM_NAME': fields['LINE_ITEM'],
                        'TYPE': None,
                        'AMOUNT': fields['AMOUNT']
                    })
                    row_count += 1
        
        # Progress indicator for large files
        if row_count % 1000 == 0:
            print(f"  Processed {row_count} rows...")
    
    wb.close()
    
    print(f"Total rows processed: {row_count}")
    
    # Create DataFrame
    df_all = pd.DataFrame(cleaned_rows)
    
    # Separate into different DataFrames
    df_projects = df_all[df_all['ROW_TYPE'] == 'PROJECT'].copy()
    df_budget_lines = df_all[df_all['ROW_TYPE'] == 'BUDGET_LINE'].copy()
    
    # Drop ROW_TYPE column from separated sheets
    df_projects = df_projects.drop('ROW_TYPE', axis=1)
    df_budget_lines = df_budget_lines.drop('ROW_TYPE', axis=1)
    
    # Rename ITEM_NAME appropriately
    df_projects = df_projects.rename(columns={'ITEM_NAME': 'PROJECT_NAME'})
    df_budget_lines = df_budget_lines.rename(columns={'ITEM_NAME': 'LINE_ITEM'})
    
    print(f"\n{'='*60}")
    print("DATA SUMMARY")
    print('='*60)
    print(f"Total rows: {len(df_all)}")
    print(f"  - Projects: {len(df_projects)}")
    print(f"  - Budget Lines: {len(df_budget_lines)}")
    print(f"  - MDA Headers: {len(df_all) - len(df_projects) - len(df_budget_lines)}")
    
    # Preview projects
    print(f"\n{'='*60}")
    print("PROJECTS PREVIEW (first 5)")
    print('='*60)
    if len(df_projects) > 0:
        print(df_projects[['MDA_CODE', 'CODE', 'PROJECT_NAME', 'TYPE', 'AMOUNT']].head())
    else:
        print("No projects found")
    
    # Check for projects with missing names or amounts
    missing_names = df_projects['PROJECT_NAME'].isna().sum()
    missing_amounts = df_projects['AMOUNT'].isna().sum()
    missing_types = df_projects['TYPE'].isna().sum()
    
    if missing_names > 0 or missing_amounts > 0:
        print(f"\n{'='*60}")
        print("DATA QUALITY CHECK")
        print('='*60)
        print(f"Projects with missing PROJECT_NAME: {missing_names}")
        print(f"Projects with missing AMOUNT: {missing_amounts}")
        print(f"Projects with missing TYPE: {missing_types} (acceptable)")
    
    # Preview budget lines
    print(f"\n{'='*60}")
    print("BUDGET LINES PREVIEW (first 10)")
    print('='*60)
    if len(df_budget_lines) > 0:
        print(df_budget_lines[['MDA_CODE', 'CODE', 'LINE_ITEM', 'AMOUNT']].head(10))
    else:
        print("No budget lines found")
    
    # Save to Excel with multiple sheets
    print(f"\n{'='*60}")
    print(f"Writing to: {output_file}")
    print('='*60)
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Projects sheet
        df_projects.to_excel(writer, sheet_name='Projects', index=False)
        print(f"  ✓ Sheet 'Projects' created ({len(df_projects)} rows)")
        
        # Budget lines sheet
        df_budget_lines.to_excel(writer, sheet_name='Budget_Lines', index=False)
        print(f"  ✓ Sheet 'Budget_Lines' created ({len(df_budget_lines)} rows)")
        
        # All data sheet
        df_all.to_excel(writer, sheet_name='All_Data', index=False)
        print(f"  ✓ Sheet 'All_Data' created ({len(df_all)} rows)")
    
    print(f"\n{'='*60}")
    print("✓ DONE!")
    print('='*60)
    print(f"Output file: {output_file}")
    print(f"Sheets: Projects, Budget_Lines, All_Data")
    print(f"\nV4 Features:")
    print(f"  - Sequential cell-level parsing (no fixed columns)")
    print(f"  - Resilient to column shifts and misalignment")
    print(f"  - Handles missing TYPE values")
    print(f"  - Reports data quality issues")
    
    return df_projects, df_budget_lines, df_all


if __name__ == "__main__":
    # ============================================================
    # EDIT THESE PATHS FOR YOUR 45MB FILE
    # ============================================================
    input_file = "data_approved/inputs/full_budget_file.xlsx"
    output_file = "data_approved/outputs/v2_consolidated_budget.xlsx"
    
    # Run the consolidation
    df_projects, df_budget_lines, df_all = consolidate_budget_excel_v4(input_file, output_file)