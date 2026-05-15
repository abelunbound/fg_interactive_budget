#!/usr/bin/env python3
"""
Consolidate spread-out Nigerian budget Excel data into clean table structure.
V3: Use CODE pattern as primary classifier (more robust than TYPE column).

USAGE:
    python3 consolidate_budget_v3.py

BEFORE RUNNING:
    Edit lines 176-177 to point to your files:
    input_file = '/path/to/your/full_budget_file.xlsx'
    output_file = '/path/to/output/consolidated_budget.xlsx'

Classification Logic:
1. CODE starts with "ERGP" → Project (TYPE optional)
2. CODE is 8+ digit number, no TYPE, no AMOUNT → MDA Header
3. Everything else with AMOUNT → Budget Line
"""

import pandas as pd
from openpyxl import load_workbook

def consolidate_budget_excel_v3(input_file, output_file):
    """
    Consolidate budget data using CODE pattern as primary classifier.
    
    Args:
        input_file: Path to input Excel file (the messy 71-column file)
        output_file: Path to output Excel file (clean multi-sheet file)
    
    Returns:
        tuple: (df_projects, df_budget_lines, df_all)
    """
    
    # The relevant column indices (0-based)
    COL_CODE = 0      # Column A
    COL_NAME = 12     # Column M
    COL_TYPE = 49     # Column AX
    COL_AMOUNT = 58   # Column BG
    
    print(f"Reading: {input_file}")
    print("This may take a few minutes for large files...")
    
    # Load workbook
    wb = load_workbook(input_file, read_only=True)
    ws = wb.active
    
    # Collect cleaned data
    cleaned_rows = []
    current_mda_code = None
    current_mda_name = None
    
    print("Processing rows...")
    row_count = 0
    
    for row_num, row in enumerate(ws.iter_rows(values_only=True), 1):
        # Extract the 4 relevant columns
        code = row[COL_CODE] if COL_CODE < len(row) else None
        name = row[COL_NAME] if COL_NAME < len(row) else None
        type_val = row[COL_TYPE] if COL_TYPE < len(row) else None
        amount = row[COL_AMOUNT] if COL_AMOUNT < len(row) else None
        
        # Skip completely empty rows
        if not any([code, name, type_val, amount]):
            continue
        
        # Skip header rows (CODE, PROJECT NAME, TYPE, AMOUNT)
        if code == "CODE" and name in ["PROJECT NAME", "LINE ITEM"]:
            continue
        
        # Convert code to string for pattern matching
        code_str = str(code).strip() if code else ""
        has_amount = amount is not None and str(amount).strip() != ""
        
        # PRIMARY CLASSIFIER: CODE PATTERN
        
        # 1. ERGP code → Always a Project
        if code_str.startswith("ERGP"):
            cleaned_rows.append({
                'ROW_TYPE': 'PROJECT',
                'MDA_CODE': current_mda_code,
                'MDA_NAME': current_mda_name,
                'CODE': code,
                'ITEM_NAME': name,
                'TYPE': type_val,
                'AMOUNT': amount
            })
            row_count += 1
        
        # 2. 8+ digit numeric code without amount → MDA Header
        elif code_str.isdigit() and len(code_str) >= 8 and not has_amount:
            current_mda_code = code
            current_mda_name = name
            
            cleaned_rows.append({
                'ROW_TYPE': 'MDA_HEADER',
                'MDA_CODE': code,
                'MDA_NAME': name,
                'CODE': None,
                'ITEM_NAME': None,
                'TYPE': None,
                'AMOUNT': None
            })
            row_count += 1
        
        # 3. Everything else with amount → Budget Line
        elif code and name and has_amount:
            cleaned_rows.append({
                'ROW_TYPE': 'BUDGET_LINE',
                'MDA_CODE': current_mda_code,
                'MDA_NAME': current_mda_name,
                'CODE': code,
                'ITEM_NAME': name,
                'TYPE': type_val,
                'AMOUNT': amount
            })
            row_count += 1
        
        # Progress indicator for large files
        if row_count % 1000 == 0:
            print(f"  Processed {row_count} rows...")
    
    wb.close()
    
    print(f"Total rows processed: {row_count}")
    
    # Create DataFrame
    df_all = pd.DataFrame(cleaned_rows)
    
    # Separate into different DataFrames
    df_projects = df_all[df_all['ROW_TYPE'] == 'PROJECT'].copy()
    df_budget_lines = df_all[df_all['ROW_TYPE'] == 'BUDGET_LINE'].copy()
    
    # Drop ROW_TYPE column from separated sheets
    df_projects = df_projects.drop('ROW_TYPE', axis=1)
    df_budget_lines = df_budget_lines.drop('ROW_TYPE', axis=1)
    
    # Rename ITEM_NAME appropriately
    df_projects = df_projects.rename(columns={'ITEM_NAME': 'PROJECT_NAME'})
    df_budget_lines = df_budget_lines.rename(columns={'ITEM_NAME': 'LINE_ITEM'})
    
    print(f"\n{'='*60}")
    print("DATA SUMMARY")
    print('='*60)
    print(f"Total rows: {len(df_all)}")
    print(f"  - Projects: {len(df_projects)}")
    print(f"  - Budget Lines: {len(df_budget_lines)}")
    print(f"  - MDA Headers: {len(df_all) - len(df_projects) - len(df_budget_lines)}")
    
    # Preview projects
    print(f"\n{'='*60}")
    print("PROJECTS PREVIEW (first 5)")
    print('='*60)
    if len(df_projects) > 0:
        print(df_projects[['MDA_CODE', 'CODE', 'PROJECT_NAME', 'TYPE', 'AMOUNT']].head())
    else:
        print("No projects found")
    
    # Preview budget lines
    print(f"\n{'='*60}")
    print("BUDGET LINES PREVIEW (first 10)")
    print('='*60)
    if len(df_budget_lines) > 0:
        print(df_budget_lines[['MDA_CODE', 'CODE', 'LINE_ITEM', 'AMOUNT']].head(10))
    else:
        print("No budget lines found")
    
    # Save to Excel with multiple sheets
    print(f"\n{'='*60}")
    print(f"Writing to: {output_file}")
    print('='*60)
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Projects sheet
        df_projects.to_excel(writer, sheet_name='Projects', index=False)
        print(f"  ✓ Sheet 'Projects' created ({len(df_projects)} rows)")
        
        # Budget lines sheet
        df_budget_lines.to_excel(writer, sheet_name='Budget_Lines', index=False)
        print(f"  ✓ Sheet 'Budget_Lines' created ({len(df_budget_lines)} rows)")
        
        # All data sheet
        df_all.to_excel(writer, sheet_name='All_Data' , index=False)
        print(f"  ✓ Sheet 'All_Data' created ({len(df_all)} rows)")
    
    print(f"\n{'='*60}")
    print("✓ DONE!")
    print('='*60)
    print(f"Output file: {output_file}")
    print(f"Sheets: Projects, Budget_Lines, All_Data")
    print(f"\nClassification logic:")
    print(f"  - ERGP* codes → Projects")
    print(f"  - 8+ digit codes (no amount) → MDA Headers")
    print(f"  - Others (with amount) → Budget Lines")
    
    return df_projects, df_budget_lines, df_all


if __name__ == "__main__":
    # ============================================================
    # FILE paths
    # ============================================================
    input_file = "data_approved/inputs/full_budget_file.xlsx"
    output_file = "data_approved/outputs/consolidated_budget.xlsx"

    # pdf_path = "data_approved/inputs/full_budget_file.xlsx"
    # output_path = "data_approved/outputs/consolidated_budget.xlsx"
    
    # Run the consolidation
    df_projects, df_budget_lines, df_all = consolidate_budget_excel_v3(input_file, output_file)